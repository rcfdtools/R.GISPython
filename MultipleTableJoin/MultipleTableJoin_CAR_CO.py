# -*- coding: UTF-8 -*-
# Create a dataset file from monthly CAR - Cundinamarca - Colombia from Excel records
# Data from https://www.car.gov.co/vercontenido/2524

# Libraries
import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd


# General parameters
input_path = 'D:/R.GISPython/MultipleTableJoin/CAR_CO/Source/'
output_path = 'D:/R.GISPython/MultipleTableJoin/CAR_CO/Joined/'
parameter_value = 'CD_PTPM_M'  # More info https://github.com/rcfdtools/R.GISPython/tree/main/MultipleTableJoin/CAR_CO
excel_file = '5de950006fde6.xlsx'
clean_file = 'Clean_'+excel_file
pivot_file = 'Pivot_'+parameter_value+'.csv'
unpivot_file = 'Unpivot_'+parameter_value+'.csv'
head_rows = 15  # Header rows to delete
index_name = 'Id'
parameter_name = 'Label'
station_code = 'Code'
year_name = 'Year'
month_name = 'Month'
day_name = 'Day'
date_name = 'Date'
value_name = 'Value'
run_clean = True  # Execute the clean Excel process
run_pivot_dataset = True  # Execute the pivot dataset creation
run_unpivot_dataset = True  # Execute the unpivot dataset creation
year_to_integer = True  # Convert Year values to integer

book = load_workbook(input_path+excel_file)
catalog = book['Catalogo']
book.remove(catalog)
sheets = book.sheetnames

# Clean headers and intermediate columns
if run_clean:
    print('\nRunning the cleaning process\n'+ 40 * '-')
    for i in sheets:
        print('Processing: %s' %i)
        sheet = book[i]
        print('\tTotal rows: %d' % sheet.max_row)
        for merge in list(sheet.merged_cells):
            print('\tUnmerging %s' % merge)
            sheet.unmerge_cells(range_string=str(merge))
        #sheet.delete_rows(idx=1)  # Delete an specific row
        sheet.delete_rows(1,head_rows)
        print('\tDeleting headers')
        print('\tDeleting white columns')
        sheet.delete_cols(2)
        for j in range(2,15):  # Delete intermediate columns
            sheet.delete_cols(j)
        print('\tRows: %d' % len(sheet['A']))
    book.save(output_path+clean_file)


# Pivot dataset creation
if run_pivot_dataset:
    print('\n\nRunning the pivot dataframe creation\n'+ 40 * '-')
    xl = pd.ExcelFile(output_path+clean_file)
    sheets = xl.sheet_names
    df = pd.DataFrame()
    for i in sheets:
        print('Processing: %s' % i)
        df1 = pd.read_excel(output_path+clean_file, sheet_name=i, header=None, names=[year_name, '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12'])
        df1[station_code] = i
        df1[parameter_name] = parameter_value
        df = pd.concat([df, df1])
    if year_to_integer:
        df = df.astype({year_name : int})
    print(df)
    df.to_csv(output_path+pivot_file, encoding='utf-8', index=False)


# Unpivot dataset creation
if run_unpivot_dataset:
    print('\n\nRunning the unpivot dataframe creation\n'+ 40 * '-')
    df = pd.read_csv(output_path+pivot_file, low_memory=False)
    print(df.info(), '\n')
    print(df.dtypes, '\n')
    print(df, '\n')
    df_unpivot = pd.melt(df, id_vars=(year_name, station_code, parameter_name), var_name=month_name, value_name=value_name, ignore_index = True)
    df_unpivot[day_name] = '01'
    df_unpivot[date_name] = pd.to_datetime(df_unpivot[[year_name, month_name, day_name]])
    df_unpivot.index.name = index_name
    df_unpivot = df_unpivot[[parameter_name, station_code, year_name, month_name, day_name, date_name, value_name]]
    print(df_unpivot, '\n')
    df_unpivot.to_csv(output_path+unpivot_file)

print('\nClean file: %s' % output_path+clean_file,
      '\nPivot file: %s' % output_path+pivot_file,
      '\nUnpivot file: %s' % output_path+unpivot_file)


# Refs
# https://blog.aspose.com/cells/insert-and-delete-rows-and-columns-in-excel-using-python/
# https://www.youtube.com/watch?v=718edSNvKLA
# https://stackoverflow.com/questions/23527887/getting-sheet-names-from-openpyxl
# https://www.geeksforgeeks.org/how-to-delete-one-or-more-rows-in-excel-using-openpyxl/
# https://www.codespeedy.com/delete-rows-of-a-sheet-using-openpyxl/
# https://stackoverflow.com/questions/69891944/unmerge-every-cell-in-an-excel-worksheet-using-openpyxl
# https://stackoverflow.com/questions/40166714/replace-specific-values-in-openpyxl
# https://stackoverflow.com/questions/17977540/pandas-looking-up-the-list-of-sheets-in-an-excel-file
# https://practicaldatascience.co.uk/data-science/how-to-reorder-pandas-dataframe-columns